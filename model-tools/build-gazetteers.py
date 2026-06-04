"""Bygg .txt-gazetteers för de deterministiska PII-detektorerna från råa öppna data.

Tar en rå export från SCB / Skolverket / Lantmäteriet och skriver rätt .txt-format
(ett namn per rad, #-header behålls) till src-tauri/src/data/. Slutanvändaren behöver
aldrig köra detta – listorna checkas in färdiga. Kör i model-tools/.venv (har openpyxl).

Källor och användning
---------------------
SCB namnstatistik (tilltalsnamn + efternamn), xlsx "namn med minst två bärare":
    python build-gazetteers.py scb-namn  _export/gazetteer/scb-namn-2022.xlsx
    -> skriver fornamn.txt och efternamn.txt

Skolverkets skolenhetsregister (skolenhetsnamn), CSV/JSON-export:
    python build-gazetteers.py skolverket  _export/gazetteer/skolenheter.csv
    -> skriver skolor.txt (bara namn som INTE redan fångas av suffix-regeln)

SCB/Lantmäteriet tätorter/orter, CSV/xlsx:
    python build-gazetteers.py tatorter  _export/gazetteer/tatorter.xlsx
    -> skriver ortnamn.txt

Matchningen i appen är skiftlägeskänslig (kräver inledande versal), så namnen
title-case:as ("ERIK" -> "Erik", "ANNA-KARIN" -> "Anna-Karin").
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "src-tauri" / "src" / "data"

# Tröskel för "vanliga" namn (antal bärare). SCB-listan har en lång svans av sällsynta
# namn som mest ökar risken för övermaskering; vi tar de vanliga. Justera vid behov.
DEFAULT_MIN_BEARERS = 500

# Tröskel för tätorter (antal invånare). Små tätorter har ofta vardagsordsnamn; börja stort.
DEFAULT_MIN_POP = 5000

# Tillåtna tecken i ett namn/ortnamn (svenska + vanliga lånebokstäver, bindestreck,
# apostrof, mellanslag). Filtrerar bort initialer ("A-C"), siffror och skräp.
VALID = re.compile(r"^[A-Za-zÀ-ÖØ-öø-ÿ][A-Za-zÀ-ÖØ-öø-ÿ\-' ]*$")


def titlecase(name: str) -> str:
    """'ERIK' -> 'Erik', 'ANNA-KARIN' -> 'Anna-Karin', 'VON ESSEN' -> 'Von Essen'.

    Versaliserar varje del avgränsad av mellanslag eller bindestreck och behåller
    avgränsaren. Använder str.capitalize() per del så å/ä/ö hanteras rätt.
    """
    out = []
    token = ""
    for ch in name:
        if ch in " -":
            out.append(token.capitalize())
            out.append(ch)
            token = ""
        else:
            token += ch
    out.append(token.capitalize())
    return "".join(out)


def looks_like_initials(name: str) -> bool:
    """True för 'A', 'A-C', 'A B' o.dyl. (varje del högst ett tecken)."""
    parts = re.split(r"[ \-]", name)
    return all(len(p) <= 1 for p in parts)


def clean_terms(pairs: list[tuple[str, int]], min_bearers: int) -> list[str]:
    """Filtrera (namn, antal) -> sorterad, deduplicerad lista title-case:ade namn."""
    seen: dict[str, None] = {}
    for raw, cnt in pairs:
        if cnt < min_bearers:
            continue
        name = str(raw).strip()
        if len(name) < 2 or looks_like_initials(name) or not VALID.match(name):
            continue
        seen[titlecase(name)] = None
    return sorted(seen)


def write_list(path: Path, header: str, terms: list[str]) -> None:
    body = header.rstrip() + "\n\n" + "\n".join(terms) + "\n"
    path.write_text(body, encoding="utf-8")
    print(f"  {path.relative_to(ROOT)}: {len(terms)} rader")


# --- SCB namn -------------------------------------------------------------------

def read_scb_sheet(wb, sheet: str) -> list[tuple[str, int]]:
    ws = wb[sheet]
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:  # rad 0–4 = rubriker/metadata, namn i kolumn A, antal i kolumn B
            continue
        name, cnt = row[0], row[1]
        if name and isinstance(cnt, (int, float)):
            out.append((str(name), int(cnt)))
    return out


def cmd_scb_namn(args: argparse.Namespace) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    forn = read_scb_sheet(wb, "Tilltalsnamn kvinnor") + read_scb_sheet(wb, "Tilltalsnamn män")
    efter = read_scb_sheet(wb, "Efternamn")

    forn_terms = clean_terms(forn, args.min_bearers)
    efter_terms = clean_terms(efter, args.min_bearers)

    write_list(DATA / "fornamn.txt", FORNAMN_HEADER.format(min=args.min_bearers), forn_terms)
    write_list(DATA / "efternamn.txt", EFTERNAMN_HEADER.format(min=args.min_bearers), efter_terms)


# --- Skolverket skolenheter -----------------------------------------------------

def read_rows(src: Path) -> list[dict]:
    """Läs en CSV (semikolon/komma) eller JSON-export som en lista av dict-rader.
    För JSON tas den första list-värdet (t.ex. {"Skolenheter": [...]}) eller en topp-lista."""
    if src.suffix.lower() == ".json":
        import json

        data = json.loads(src.read_text(encoding="utf-8"))
        if isinstance(data, list):
            return data
        return next((v for v in data.values() if isinstance(v, list)), [])
    import csv

    text = src.read_text(encoding="utf-8-sig")
    first = text.splitlines()[0]
    delim = ";" if first.count(";") >= first.count(",") else ","
    return list(csv.DictReader(text.splitlines(), delimiter=delim))


def field(row: dict, candidates: tuple[str, ...]) -> str | None:
    low = {k.lower(): k for k in row}
    for c in candidates:
        k = low.get(c.lower())
        if k and row.get(k):
            return str(row[k])
    return None


# Generiska skolord. Ett namn som BARA består av sådana ("Anpassad grundskola", "Kommunal
# vuxenutbildning", "Vuxenutbildningen SFI") är en allmän term, inte ett egennamn, och skulle
# övermaska vanlig elevhälsotext. Vi kräver minst ett särskiljande (icke-generiskt) ord.
GENERIC_SCHOOL_WORDS = {
    "anpassad", "anpassade", "kommunal", "kommunala", "fristående", "grundläggande", "gymnasial",
    "vuxen", "vuxna", "för", "och", "i", "som", "vid", "av", "med", "samt", "den", "det",
    "skola", "skolan", "grundskola", "grundskolan", "grundsärskola", "grundsärskolan",
    "gymnasieskola", "gymnasieskolan", "gymnasium", "gymnasiet", "förskola", "förskolan",
    "särskola", "särskolan", "särskild", "särskilda", "friskola", "fritidshem", "resursskola",
    "resursskolan", "vuxenutbildning", "vuxenutbildningen", "komvux", "särvux", "lärvux",
    "sfi", "vux", "im", "ab", "utbildning", "utbildningen", "yrkesutbildning", "yrkesvux", "yrkes",
    "introduktion", "introduktionsprogram", "introduktionsprogrammet", "inriktning", "ämne",
    "ämnesområde", "undervisning", "undervisningsgrupp", "nivå", "anordnare", "externa", "extern",
    "enhet", "enheten",
}


def has_distinctive_word(name: str) -> bool:
    """True om namnet har minst ett ord som inte är ett generiskt skolord (ett egennamn)."""
    return any(t.strip("-").lower() not in GENERIC_SCHOOL_WORDS and len(t) >= 2 for t in name.split())


def cmd_skolverket(args: argparse.Namespace) -> None:
    rows = read_rows(args.file)
    seen: dict[str, None] = {}
    for r in rows:
        status = field(r, ("Status", "status"))
        if status and status.lower() != "aktiv":
            continue  # hoppa över vilande/planerade/nedlagda enheter
        name = field(r, ("Skolenhetsnamn", "skolenhetsnamn", "namn", "name"))
        if not name:
            continue
        name = name.strip()
        # Behåll bara FLERORDS-namn (innehåller mellanslag). Enordsnamn är antingen redan
        # fångade av rules::skolnamn ("Björkskolan", "Kunskapsskolan") eller generiska ord med
        # hög risk för övermaskering ("Elevhälsan", "Introduktionsprogrammet", "Borgen").
        # Flerordsnamn ("Internationella Engelska Skolan Bromma", "IES Bromma") är specifika.
        if " " not in name or len(name) < 2 or not VALID.match(name):
            continue
        if not has_distinctive_word(name):
            continue  # bara generiska skolord ("Anpassad grundskola") -> hoppa över
        seen[name] = None  # skrivs som de står (egennamn, redan korrekt versaliserade)
    write_list(DATA / "skolor.txt", SKOLOR_HEADER, sorted(seen))


# --- Tätorter / ortnamn ---------------------------------------------------------

# Tätortsnamn som också är vanliga svenska ord (eller vanliga förnamn) – uteslut ur
# ortnamn.txt. Platser får hög score, så en vardagsordskrock ("Bro", "Vi") maskas i varje
# dokument; namnkrockar ("Åsa") täcks redan av förnamnslistan. Justera vid behov.
STOPWORDS_ORT = {
    "Bro", "Vi", "Viken", "Bo", "By", "Ed", "Lo", "Näs", "Vik", "Ås", "Hed", "Hov", "Backe",
    "Sand", "Mon", "Åsa",
}


def load_kommuner() -> set[str]:
    """Kommunnamn ur kommuner.txt – tätorter som delar namn med en kommun (oftast
    centralorten) utelämnas ur ortnamn.txt, de täcks redan av kommun-gazetteern."""
    path = DATA / "kommuner.txt"
    if not path.exists():
        return set()
    return {l.strip() for l in path.read_text(encoding="utf-8").splitlines() if l.strip() and not l.startswith("#")}


def to_int(v) -> int | None:
    try:
        return int(str(v).replace(" ", "").replace("\xa0", ""))
    except (ValueError, TypeError):
        return None


def cmd_tatorter(args: argparse.Namespace) -> None:
    src = args.file
    pairs: list[tuple[object, object]] = []  # (namn, folkmängd|None)
    if src.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl

        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        name_col = pop_col = header_idx = None
        for i, r in enumerate(rows[:15]):
            for j, cell in enumerate(r):
                if not isinstance(cell, str):
                    continue
                if name_col is None and re.search(r"tätort|ortnamn|\bort\b|namn", cell, re.I):
                    header_idx, name_col = i, j
                if re.search(r"folkmängd|befolkning|invånare", cell, re.I):
                    pop_col = j
            if name_col is not None:
                break
        if name_col is None:
            sys.exit("hittade ingen tätorts-/namnkolumn i xlsx")
        for r in rows[header_idx + 1:]:
            pop = r[pop_col] if pop_col is not None and pop_col < len(r) else None
            pairs.append((r[name_col], pop))
    else:
        for r in read_rows(src):
            name = field(r, ("Tätort", "tätort", "ortnamn", "namn", "name"))
            pop = field(r, ("Folkmängd", "folkmängd", "folkmangd", "befolkning", "invånare", "population"))
            pairs.append((name, pop))

    kommuner = load_kommuner()
    seen: dict[str, None] = {}
    for name, pop in pairs:
        if not name:
            continue
        p = to_int(pop)
        if p is not None and p < args.min_pop:
            continue  # för liten tätort – hög risk för vardagsordskrock, hoppa över
        # Dela ihopslagna tätorter ("Upplands Väsby och Sollentuna", "Skanör med Falsterbo").
        for part in re.split(r"\s+(?:och|med)\s+", str(name)):
            part = part.strip()
            if len(part) < 2 or looks_like_initials(part) or not VALID.match(part):
                continue
            if part in kommuner or part in STOPWORDS_ORT:
                continue  # täcks redan av kommuner.txt, eller vardagsord/förnamn
            seen[part] = None  # SCB:s tätortsnamn är redan korrekt versaliserade
    write_list(DATA / "ortnamn.txt", ORTNAMN_HEADER.format(min=args.min_pop), sorted(seen))


# --- Headers (behålls överst i varje fil) ---------------------------------------

FORNAMN_HEADER = """\
# Svenska förnamn (tilltalsnamn) – gazetteer för Category::Person.
# Källa: SCB:s namnstatistik, "tilltalsnamn med minst två bärare" (31 dec 2022).
# Genererad av model-tools/build-gazetteers.py (tröskel: minst {min} bärare).
#
# Format: ett namn per rad. Rader som börjar med # och tomma rader ignoreras.
# Matchningen är skiftlägeskänslig (kräver inledande versal) eftersom många namn
# även är vardagsord (Björn, Sten, My, Ros). Träffar får låg score och granskas."""

EFTERNAMN_HEADER = """\
# Svenska efternamn – gazetteer för Category::Person.
# Källa: SCB:s namnstatistik, "efternamn med minst två bärare" (31 dec 2022).
# Genererad av model-tools/build-gazetteers.py (tröskel: minst {min} bärare).
#
# Format: ett namn per rad. #-rader och tomma rader ignoreras. Matchningen är
# skiftlägeskänslig (kräver inledande versal). Träffar får låg score och granskas."""

SKOLOR_HEADER = """\
# Skolnamn (flerordsnamn på aktiva skolenheter) – gazetteer för Category::Plats.
# Källa: Skolverkets skolenhetsregister (öppna API, v1/skolenhet).
# Genererad av model-tools/build-gazetteers.py.
#
# Bara FLERORDS-namn tas med ("Internationella Engelska Skolan Bromma", "IES Bromma").
# Enordsnamn utelämnas: de fångas antingen redan av rules::skolnamn ("Björkskolan",
# "Kunskapsskolan") eller är generiska ord med hög risk för övermaskering
# ("Elevhälsan", "Introduktionsprogrammet").
#
# Format: ett namn per rad. #-rader och tomma rader ignoreras. Skiftlägeskänslig."""

ORTNAMN_HEADER = """\
# Ortnamn (tätorter) – gazetteer för Category::Plats.
# Källa: SCB, "Folkmängd i tätorter per tätort" (2023, MI0810). Tätorter som delar
# namn med en kommun är utelämnade (täcks av kommuner.txt).
# Genererad av model-tools/build-gazetteers.py (tröskel: minst {min} invånare).
#
# OBS: ju större lista (lägre tröskel), desto högre risk för övermaskering av
# vardagsord. Börja smått. #-rader/tomma rader ignoreras. Matchningen är
# skiftlägeskänslig (kräver inledande versal)."""

COMMANDS = {
    "scb-namn": cmd_scb_namn,
    "skolverket": cmd_skolverket,
    "tatorter": cmd_tatorter,
}


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("source", choices=COMMANDS, help="vilken källa exporten kommer från")
    p.add_argument("file", type=Path, help="sökväg till den råa exporten (xlsx/csv/json)")
    p.add_argument("--min-bearers", type=int, default=DEFAULT_MIN_BEARERS,
                   help=f"minsta antal bärare för SCB-namn (default {DEFAULT_MIN_BEARERS})")
    p.add_argument("--min-pop", type=int, default=DEFAULT_MIN_POP,
                   help=f"minsta antal invånare för tätorter (default {DEFAULT_MIN_POP})")
    args = p.parse_args()
    if not args.file.exists():
        sys.exit(f"filen finns inte: {args.file}")
    print(f"Bygger gazetteer från {args.file.name} ({args.source})…")
    COMMANDS[args.source](args)
    print("Klart.")


if __name__ == "__main__":
    main()
